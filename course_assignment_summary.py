#
#       course_assignment_summary.py
#
#       This code will create an Excel workbook using openpyxl which will
#       contain summary details of courses and assignments for the users account
#
#       This script will also send feedback reminders to relevant memebers of
#       staff if assignments are left ungraded after a set number of working days
#       from the assignment due date
#
#       Things that need to be set:
#
#       output_dir - path of Excel file to be saved
output_dir = '/mnt/metadmin/CANVASBOTS/'
#
#       name_extenstion - name of Excel file after Canvas course number
name_extension = 'assignment_summary'
#
#
#       included_terms - list of terms that the user wishes to recieve when
#       running the code, format = 'yyyy/yy' or 'Default term'
included_terms = ['2022/23']
#
#
#       offset_days - working days after which the due date of an assignment
#       should be
offset_days = 15
#
#
#       reminder_dates - days from due date for reminder emails to be sent
reminder_dates = [
	[10], # day to email reminder
	[14], # day to email final warning
	[16] # day after deadline
	]
#
#
#       TSO_email - email of TSO
TSO_email = ['metmat-eso@contacts.bham.ac.uk', 'M.J.JENKINS@bham.ac.uk']
#
#
#
import sys
sys.path.append("../PyCAPI/module/") # First two lines are needed for import of PyCAPI
import PyCAPI
import uob_utils
import datetime
import json
import numpy as np
import pandas
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from openpyxl import worksheet
import re
from copy import deepcopy


capi = PyCAPI.CanvasAPI()
mail = uob_utils.MailAPI()
today = datetime.datetime.now()



###############################################################################
# Create workbook for storing information in
#
wb = Workbook()
wb.remove(wb.active) # Remove initially created sheet
ws = wb.create_sheet(title='Courses') # Set active worksheet and name



###############################################################################
# Set columns for course summary sheet (this way one needs to do this once)
(	col_term, col_cid, col_cnm, col_ccd, col_cadmin, col_cavail, col_tenroll, 
	col_senroll
 ) = [chr(65 + i) for i in range(8)] # A -> H assignment
# Format course summary sheet and column widths
cols = { # key = column name, col_xyz = column reference, int = column width
	'Term' : [col_term, 10], 'Course ID' : [col_cid, 11], 
	'Course Name' : [col_cnm, 30], 'Course Code' : [col_ccd, 15],
	'Admin' : [col_cadmin, 15], 'Availability' : [col_cavail, 15],
	'Teachers Enrolled' : [col_tenroll, 20], 
	'Students Enrolled' : [col_senroll, 20]
	}
# Format course summary sheet and column widths
for key in cols:
	ws[cols[key][0] + '1'] = key
	ws.column_dimensions[cols[key][0]].width = cols[key][1]


###############################################################################
# Retrieve data of all courses from Canvas account
#
# Following 4 lines are needed to return courses that only I am a teacher on
#payload = {}
#payload['enrollment_type'] = 'teacher' # Remove this line if used by admin
#payload['include[]'] = ['term', 'teachers', 'total_students']
#courses = capi.get('/courses', payload=payload) # Remove this line if used by admin
#courses = capi.get_courses(payload=payload) # Use this line when used by Met&Mat admin
#print json.dumps(courses, indent = 2)
#allcourses = capi.get('/courses', payload=payload)
allcourses = capi.get_courses(account_id='114', include=['term', 'teachers', 'total_students'])
courses = []
for course in allcourses:
	if course['term']['name'] in included_terms:
		courses.append(course)
#print json.dumps(courses, indent = 2)


i = 2 # row in ws
for course in courses: # Loop through all courses in Canvas account
	ws[col_term+str(i)] = course['term']['name']
	ws[col_cid+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['id'])+'")'
	ws[col_cnm+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['name'])+'")'
	ws[col_ccd+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['course_code'])+'")'
	ws[col_cadmin+str(i)] = course['account_id']
	ws[col_cavail+str(i)] = course['workflow_state'].title()
	if course['workflow_state'] != 'available':
		for cell in ws[str(i)+':'+str(i)]:
			cell.font = Font(color='00D3D3D3')
	ws[col_tenroll+str(i)] = len(course['teachers'])
	ws[col_senroll+str(i)] = course['total_students']
	i += 1

# Make the workbook into a filtered table.
ws.auto_filter.ref = 'A1:AA'+str(i)



###############################################################################
# Summarise assignments in a different Excel sheet
#
ws = wb.create_sheet(title='Assignments') # Set active worksheet



###############################################################################
# Set columns for assignment summary sheet (this way one needs to do this once)
#
(	col_cid, col_cnm, col_ccd, col_term, col_asgnid, col_asgnnm, col_cadmin,
	col_cavail, col_asgnavail, col_asgntype, col_group, col_unlock, col_due,
	col_lock, col_sub, col_ungr, col_miss, col_late, col_median, col_points,
	col_grby, col_fingr, col_weight, col_muted, col_manual, col_daysafter
) = [chr(65 + i) for i in range(26)] # A -> Z assignment (will not work if more cols added!)
cols = { # key = column name, col_xyz = column reference, int = column width
	'Course ID' : [col_cid, 11], 'Course Name' : [col_cnm, 20], 
	'Course Code'  : [col_ccd, 15], 'Term' : [col_term, 14],
	'Assignment ID' : [col_asgnid, 30], 'Assignment Name' : [col_asgnnm, 15], 
	'Admin' : [col_cadmin, 17], 'Course Availability' : [col_cavail, 17], 
	'Assignment Availability' : [col_asgnavail, 17], 
	'Assignment Type' : [col_asgntype, 16], 'Grouped' : [col_group, 17], 
	'Unlock at' : [col_unlock, 17], 'Due at' : [col_due, 17], 
	'Lock at' : [col_lock, 17], 'Submissions' : [col_sub, 12], 
	'Ungraded' : [col_ungr, 12], 'Missing' : [col_miss, 12], 
	'Late' : [col_late, 12], 'Median' : [col_median, 12], 
	'Points possible' : [col_points, 12], 'Grade by' : [col_grby, 12],
	'Summative or Formative' : [col_fingr, 12],	'Weighting' : [col_weight, 12],
	'Muted' : [col_muted, 15], 'Manual' : [col_manual, 15],
	'Days since Deadline' : [col_daysafter, 18]
	}
# Format assignment summary sheet and column widths
for key in cols:
	ws[cols[key][0] + '1'] = key
	ws.column_dimensions[cols[key][0]].width = cols[key][1]


###############################################################################
# Retrieve data of all assignments from Canvas account
#
i = 2
for course in courses:
	course_assignments = capi.get_assignments(course['id'])
	payload = {}
	payload['include[]'] = ['assignments']
	assignment_groups = capi.get('/courses/%s/assignment_groups' % course['id'], payload=payload)
	analytics_missing = False
	try:
		analytics = capi.get('/courses/%s/analytics/assignments' % course['id'])
	except:
		analytics_missing = True
	#print json.dumps(course_assignments, indent = 2)
	#print json.dumps(assignment_groups, indent = 2)
	for assignment in course_assignments:
		ws[col_cid+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['id'])+'")'
		ws[col_cnm+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['name'])+'")'
		ws[col_ccd+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'", "'+str(course['course_code'])+'")'  
		ws[col_term+str(i)] = course['term']['name']
		ws[col_asgnid+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'/assignments/'+str(assignment['id'])+'", "'+str(assignment['id'])+'")'
		ws[col_asgnnm+str(i)] = '=HYPERLINK("http://147.188.152.33:8080/eval.php?url=https://canvas.bham.ac.uk/courses/'+str(course['id'])+'/assignments/'+str(assignment['id'])+'", "'+str(assignment['name'].encode('ascii', 'ignore'))+'")'
		ws[col_cadmin+str(i)] = course['account_id']
		ws[col_cavail+str(i)]= course['workflow_state'].title()
		if assignment['published']:
			ws[col_asgnavail+str(i)] = 'Published'
		else:
			ws[col_asgnavail+str(i)] = 'Unpublished'
#		if 'online_upload' in assignment['submission_types'] or 'online_quiz' in assignment['submission_types'] or 'external_tool' in assignment['submission_types']:
#			ws[col_asgntype+str(i)] = 'Online'
#		else:
#			ws[col_asgntype+str(i)] = 'Paper'
		if len(assignment['submission_types']) == 1:
			ws[col_asgntype+str(i)] = assignment['submission_types'][0]
		elif len(assignment['submission_types']) == 0:
			ws[col_asgntype+str(i)] = 'None'
		else:
			ws[col_asgntype+str(i)] = 'Multiple types'
		if assignment['group_category_id'] == None:
			ws[col_group+str(i)] = 'No'
		else:
			ws[col_group+str(i)] = 'Yes'
			
		
		if analytics_missing == False:
			for analytic in analytics:
				if analytic['assignment_id'] == assignment['id']:
					assignment['analytics'] = analytic
					if analytic['tardiness_breakdown']['missing'] == None:
						ws[col_miss+str(i)] = 'None'
					else:
						ws[col_miss+str(i)] = analytic['tardiness_breakdown']['missing'] * analytic['tardiness_breakdown']['total']
					if analytic['tardiness_breakdown']['late'] == None:
						ws[col_late+str(i)] = 'None'
					else:
						ws[col_late+str(i)] = analytic['tardiness_breakdown']['late'] * analytic['tardiness_breakdown']['total']
					if analytic['tardiness_breakdown']['total'] == None:
						ws[col_sub+str(i)] = 'None'
					else:
						ws[col_sub+str(i)] = analytic['tardiness_breakdown']['total']
					if analytic['points_possible'] == None:
						ws[col_median+str(i)] = 'None'
					else:
						if analytic['median'] == None:
							ws[col_median+str(i)] = 'None'
						else:
							ws[col_median+str(i)] = analytic['median']
		else:
			ws[col_miss+str(i)] = 'Missing'
			ws[col_late+str(i)] = 'Missing'
			ws[col_median+str(i)] = 'Missing'
		
		
		if assignment['points_possible'] == None:
			ws[col_points+str(i)] = 'None set'
		else:
			ws[col_points+str(i)] = assignment['points_possible']
		
		if assignment['unlock_at'] == None:
			ws[col_unlock+str(i)] = 'None set'
		else:
			ws[col_unlock+str(i)] = str(datetime.datetime.strptime(assignment['unlock_at'], "%Y-%m-%dT%H:%M:%SZ"))
		if assignment['due_at'] == None:
			ws[col_due+str(i)] = 'None set'
			ws[col_grby+str(i)] = 'None set'
		else:
			due_date = datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ")
			ws[col_due+str(i)] = str(due_date)
			days_after, grade_by, is_working_day = uob_utils.days_since_deadline(due_date)
			#recipients = []
			#cc_recipients = []
			ws[col_grby+str(i)] = str(grade_by)	
			ws[col_daysafter + str(i)] = str(days_after)
			if (days_after in reminder_dates[0]) and assignment['needs_grading_count'] != 0 and is_working_day:
				uob_utils.produce_email('5 days left', assignment, TSO_email, ws, col_sub, i)
			elif (days_after in reminder_dates[1]) and assignment['needs_grading_count'] != 0 and is_working_day:
				uob_utils.produce_email('1 day left', assignment, TSO_email, ws, col_sub, i)
			elif days_after in reminder_dates[2] and is_working_day:
				uob_utils.produce_email('overdue', assignment, TSO_email, ws, col_sub, i)
		
		if assignment['lock_at'] == None:
			ws[col_lock+str(i)] = 'None set'
		else:
			ws[col_lock+str(i)] = str(datetime.datetime.strptime(assignment['lock_at'], "%Y-%m-%dT%H:%M:%SZ"))
		ws[col_ungr+str(i)] = assignment['needs_grading_count']

		if assignment['omit_from_final_grade'] == False:
			ws[col_fingr+str(i)] = 'Summative'
		else:
			ws[col_fingr+str(i)] = 'Formative'
		for assignment_group in assignment_groups:
			for assignment_group_assignment in assignment_group['assignments']:
				if assignment_group_assignment['id'] == assignment['id']:
					ws[col_weight+str(i)] = assignment_group['group_weight']/len(assignment_group['assignments'])
		if assignment['muted'] == False:
			ws[col_muted+str(i)] = 'No'
		else:
			ws[col_muted+str(i)] = 'Yes'
			
		if assignment['post_manually'] == False:
			ws[col_manual+str(i)] = 'No'
		else:
			ws[col_manual+str(i)] = 'Yes'
		
		# Colour the relevant lines
		if ws[col_asgntype+str(i)].value == 'Paper':
			for cell in ws[str(i)+':'+str(i)]:
				cell.font = Font(color='FF0000')
		if assignment['due_at'] != None:
			if datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(days=14) < today and assignment['needs_grading_count'] > 0:
				for cell in ws[str(i)+':'+str(i)]:
					cell.fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
			if datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ") < today and assignment['needs_grading_count'] == 0 and assignment['muted'] == False and ws[col_asgntype+str(i)].value != 'Paper':
				for cell in ws[str(i)+':'+str(i)]:
					cell.font = Font(color='00FF00')
		if course['workflow_state'] != 'available' or not assignment['published']:
			for cell in ws[str(i)+':'+str(i)]:
				cell.font = Font(color='00D3D3D3')
		
		i += 1

# Make the workbook into a filtered table.
ws.auto_filter.ref = 'A1:AA'+str(i)


###############################################################################
# Save the created Excel file with customised filename
#
#filename = name_extenstion+'_'+str(datetime.date.today())
filename = name_extension
file_extension = '.xlsx'
wb.save(filename = output_dir+filename+file_extension)
#print 'Workbook: ' + filename + ' saved'
