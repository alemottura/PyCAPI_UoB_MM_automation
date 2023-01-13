
import sys
sys.path.append("../PyCAPI/module/") # First two lines are needed for import of PyCAPI
import PyCAPI
import uob_utils
import datetime
import json
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl import Workbook

course_list = '/mnt/metadmin/CANVASBOTS/late_penalty_course_list.xlsx'

capi = PyCAPI.CanvasAPI()

try:
	wb = load_workbook(course_list , read_only = True, data_only = True)
	ws = wb['Sheet1']
except:
	raise RuntimeError('Could not open Excel file containing tutees.')
i = 2
course_list = []
assignment_list = []
while ws['A'+str(i)].value != None:
    assignments = capi.get_assignments(int(ws['A'+str(i)].value))
    for assignment in assignments:
        #print assignment
        if assignment['published'] == True:
            if assignment['due_at'] != None:
                assignment_due_date = datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ")
                now = datetime.datetime.now()
                if (now-assignment_due_date).days > 0:    
                    assignment_list.append([int(ws['A'+str(i)].value), assignment['id']])
    course_list.append(int(ws['A'+str(i)].value))
    i = i + 1



"""
def days_late(deadline, submission):
    if (submission-deadline).total_seconds() < 600:
        return 0
    if uob_utils.UniversityWeek(submission.date()) == uob_utils.UniversityWeek(deadline.date()):
        # Submission and deadline in the same week
        if submission.weekday() >= 5:
            # submitted over the weekend
            friday = datetime.datetime.combine(uob_utils.DateFromUniversityWeek(uob_utils.AcademicYear(submission.date()),uob_utils.UniversityWeek(submission.date()),4), datetime.time(23, 59, 59))
            return math.ceil((friday-deadline).total_seconds()/86400)
        else:
            return math.ceil((submission-deadline).total_seconds()/86400)
    elif uob_utils.UniversityWeek(submission.date()) == uob_utils.UniversityWeek(deadline.date())+1:
        # Submission in the following week
        if submission.weekday() >= 5:
            # submitted over the weekend
            friday = datetime.datetime.combine(uob_utils.DateFromUniversityWeek(uob_utils.AcademicYear(submission.date()),uob_utils.UniversityWeek(submission.date()),4), datetime.time(23, 59, 59))
            return math.ceil(((friday-deadline).total_seconds() - 172800)/86400)
        else:
            return math.ceil(((submission-deadline).total_seconds() - 172800)/86400)
    elif uob_utils.UniversityWeek(submission.date()) == uob_utils.UniversityWeek(deadline.date())+2:
        # Submission in the week after following week
        if submission.weekday() >= 5:
            # submitted over the weekend
            friday = datetime.datetime.combine(uob_utils.DateFromUniversityWeek(uob_utils.AcademicYear(submission.date()),uob_utils.UniversityWeek(submission.date()),4), datetime.time(23, 59, 59))
            return math.ceil(((friday-deadline).total_seconds() - 345600)/86400)
        else:
            return math.ceil(((submission-deadline).total_seconds() - 345600)/86400)
    elif uob_utils.UniversityWeek(submission.date()) == uob_utils.UniversityWeek(deadline.date())+3:
        # Submission even later...
        if submission.weekday() >= 5:
            # submitted over the weekend
            friday = datetime.datetime.combine(uob_utils.DateFromUniversityWeek(uob_utils.AcademicYear(submission.date()),uob_utils.UniversityWeek(submission.date()),4), datetime.time(23, 59, 59))
            return math.ceil(((friday-deadline).total_seconds() - 518400)/86400)
        else:
            return math.ceil(((submission-deadline).total_seconds() - 518400)/86400)

"""    
holidays=['2022-12-19','2022-12-20','2022-12-21','2022-12-22','2022-12-23','2022-12-26','2022-12-27','2022-12-28','2022-12-29','2022-12-30','2023-01-02']

def days_late(deadline, submission):
    if (submission-deadline).total_seconds() < 600:
        return 0
    no_days_late = np.busday_count(deadline.strftime("%Y-%m-%d"), submission.strftime("%Y-%m-%d"), weekmask='1111100', holidays=holidays)
    if submission.time() > deadline.time() and submission.strftime("%Y-%m-%d") not in holidays:
        no_days_late = no_days_late+1
    return no_days_late


for assignment in assignment_list:
	course_id = assignment[0]
	assignment_id = assignment[1]
	assignment = capi.get_assignment(course_id, assignment_id, include='overrides')
	personal_due_date = {}
	for override in assignment['overrides']:
#		print override['due_at']
		if 'student_ids' in override.keys():
			if len(override['student_ids']) > 0:
				for student_id in override['student_ids']:
					if override['due_at'] != None:
						personal_due_date[student_id] = datetime.datetime.strptime(override['due_at'], "%Y-%m-%dT%H:%M:%SZ")
#	print personal_due_date
	if assignment['due_at'] != None:
		assignment_due_date = datetime.datetime.strptime(assignment['due_at'], "%Y-%m-%dT%H:%M:%SZ")
	else:
		continue


	submissions = capi.get("/courses/%s/assignments/%s/submissions?grouped=false" % (course_id, assignment_id))


	
	for submission in submissions:
		if submission['submitted_at'] != None:
			submitted_at = datetime.datetime.strptime(submission['submitted_at'], "%Y-%m-%dT%H:%M:%SZ")
			if submission['user_id'] in  personal_due_date.keys():
				due_date = personal_due_date[submission['user_id']]
			else:
				due_date = assignment_due_date
			seconds_late = (submitted_at-due_date).total_seconds()
			my_delay = days_late(due_date, submitted_at)
			if seconds_late > 0:
				if my_delay > 0:        
#					print "LATE", submission['user_id'], due_date, submitted_at, str(seconds_late), str(my_delay)
#					print ' '
					payload = {}
					payload['submission[late_policy_status]'] = 'late'
					payload['submission[seconds_late_override]'] = my_delay * 86400
					capi.put('/courses/%s/assignments/%s/submissions/%s' % (course_id, assignment_id, submission['user_id']), payload=payload)
				else:
#					print "LATE BUT NOT LATE", submission['user_id'], due_date, submitted_at, str(seconds_late), str(my_delay)
					payload = {}
					payload['submission[late_policy_status]'] = 'none'
					capi.put('/courses/%s/assignments/%s/submissions/%s' % (course_id, assignment_id, submission['user_id']), payload=payload)
#					print ' '
			else:
#				print "NOT LATE", submission['user_id']
				payload = {}
				payload['submission[late_policy_status]'] = 'none'
				capi.put('/courses/%s/assignments/%s/submissions/%s' % (course_id, assignment_id, submission['user_id']), payload=payload)
#				print ' '
#		else:
#			print "NOT SUBMITTED", submission['user_id']
#			print ' '



for course in course_list:
	course_id = course
	
	payload = {}
	payload['late_policy[late_submission_deduction_enabled]'] = True
	payload['late_policy[late_submission_deduction]'] = 5
	payload['late_policy[late_submission_interval]'] = 'day'
	payload['late_policy[late_submission_minimum_percent_enabled]'] = True
	payload['late_policy[late_submission_minimum_percent]'] = 0
	payload['late_policy[missing_submission_deduction]'] = 100
	payload['late_policy[missing_submission_deduction_enabled]'] = True
	
	try:
		capi.post('/courses/%s/late_policy' % course_id, payload=payload)
	except:
		capi.patch('/courses/%s/late_policy' % course_id, payload=payload)
	
#	result = capi.get('/courses/%s/late_policy' % course_id, single=True)
	
#	print result
#

#https://canvas.bham.ac.uk:443/api/v1/courses/24191/assignments/89388/submissions/60415

#submission[late_policy_status]='late' or 'none'
#submission[seconds_late_override]=seconds










